from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

# workbook = Workbook()
# workSheet = workbook.active
#
# file_name = "bbc_news.xlsx"
#
# workSheet.title = "BBC"
#
# workSheet.cell(column=1, row=1, value="Date")
# workSheet.cell(column=2, row=1, value="Author")
# workSheet.cell(column=3, row=1, value="Main_Topics")
# workSheet.cell(column=4, row=1, value="Topic")
# workSheet.cell(column=5, row=1, value="Text")
#
# workbook.save(file_name)

# columns_my = ["Date", "Author", "Main_Topics", "Topic", "Text"]
#
# df_news = pd.DataFrame([list("zxcvb")], columns=columns_my)
# print(f"new:\n{df_news}")
#
# row = list('abcde')
# df_current = pd.DataFrame([row], columns=columns_my)
# print(f"current:\n{df_current}")
#
#
# df_news = df_news.append(df_current, ignore_index=True)
# print(f"new:\n{df_news}")

file_name = "panddas.xlsx"
new_file_name = "panddas1.xlsx"
df_exel = pd.read_excel(file_name, index_col=0)

# df_csv = pd.read_csv(file_name)

df_exel.to_excel(new_file_name)