from selenium import webdriver
from openpyxl import Workbook, load_workbook

driver = webdriver.Chrome(executable_path='/home/bateiko/Downloads/chromedriver_linux64/chromedriver')
driver.implicitly_wait(5)
driver.maximize_window()

current_url = 'https://www.bbc.com/ukrainian/topics/c4794229-7f87-43ce-ac0a-6cfcd6d3cef2'


def read_news_from_current_url(current_url):

    driver.get(current_url)

    # ------work with table ------
    name_file = "bbc.xlsx"
    sheet_name = "BBC"
    workbook = load_workbook(filename=name_file)
    workSheet = workbook[sheet_name]

    # ------- end work with table -------
    # enter to new

    list_of_news_of_Ukraine = driver.find_elements_by_xpath('//div[@class="eagle"]/div')
    len_of_news = len(list_of_news_of_Ukraine)
    current_web_elem = list_of_news_of_Ukraine[0]

    main_topics = driver.find_element_by_class_name("page-title").text

    for i in range(0, len_of_news):

        driver.get(current_url)
        current_web_elem = driver.find_elements_by_xpath('//div[@class="eagle"]/div')[i]

        print(f"is_enable:\t{current_web_elem.is_enabled()}")
        current_web_elem.click()

        date_web_element = driver.find_element_by_xpath('//li[@class="mini-info-list__item"]/div')

        date = date_web_element.text
        topic = driver.find_element_by_tag_name('h1').text
        try:
            author = driver.find_element_by_class_name('byline__name').text
        except:
            author = ''

        text_all_webElem = driver.find_elements_by_class_name("story-body__inner")

        all_text = ''
        count_paragraph = 0
        for paragraph in text_all_webElem:
            try:
                text_paragraph = paragraph.text + '\n'
                all_text += text_paragraph
                count_paragraph += 1
            except:
                continue

        workSheet.cell(column=1, row=i + 2, value=date)
        workSheet.cell(column=2, row=i + 2, value=author)
        workSheet.cell(column=3, row=i + 2, value=main_topics)
        workSheet.cell(column=4, row=i + 2, value=topic)
        workSheet.cell(column=5, row=i + 2, value=all_text)

workbook.save(filename=name_file)
driver.quit()