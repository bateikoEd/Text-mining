from selenium import webdriver
from openpyxl import Workbook, load_workbook

#create a new Firefox session
driver = webdriver.Chrome(executable_path='/home/bateiko/Downloads/chromedriver_linux64/chromedriver')
driver.implicitly_wait(5)
driver.maximize_window()

#Navigate to the aplication home page
driver.get("https://expertsystem.com/10-text-mining-examples/")

"""Після кожного заголовка іти некс поки не зустрінеться новий заговолок зчитуємо текс з тегів p"""
#get the search textbox
field_with_topics = driver.find_elements_by_xpath('//div[@class="l-section-h i-cf"]/*')
print(f'topic:\t{field_with_topics[2].text}\nattribute:\t'
      f'{field_with_topics[2].tag_name}')

flag = False
topics = []
text_under_topic = ''
text_after_topics = {}
last_topic = ''
for elem in field_with_topics:
    if elem.tag_name == 'h5':
        topics.append(elem.text)
        flag = True

        text_after_topics[last_topic] = text_under_topic
        text_under_topic = ''
        last_topic = elem.text

    if flag and elem.tag_name != 'h5':
        text_under_topic += elem.text + '\n'

file_name = 'topics_with_text.txt'

'''with open(file_name, 'w+') as f:
    for topic, text in text_after_topics.items():
        f.write(f'{topic}\n')
        f.write(text)'''

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Text mining'

worksheet.cell(row=1, column=1).value = "Topics"
worksheet.cell(row=1, column=2).value = "Text"

count_row = 2
for topic, text in text_after_topics.items():
    worksheet.cell(row=count_row, column=1).value = topic.rstrip()
    worksheet.cell(row=count_row, column=2).value = text.rstrip()
    count_row += 1

workbook.save('Text.xlsx')

driver.quit()